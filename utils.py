import pandas as pd
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from sqlalchemy import func
from model import Purchase, Transfer , PurchaseItem
from sqlalchemy.orm import Session
from openpyxl.utils import get_column_letter



def read_excel_file(file_path: str, **kwargs) -> pd.DataFrame:
    """
    Read the Excel file into a DataFrame.
    """
    return pd.read_excel(file_path, **kwargs)

def split_dataframe_in_batches(df: pd.DataFrame, batch_size: int) -> List[pd.DataFrame]:
    """
    Split DataFrame into list of DataFrames with batch_size rows each.
    """
    return [df[i:i + batch_size] for i in range(0, len(df), batch_size)]



def generate_excel_report(task_id, session, purchase_ids: List[int], transfer_ids: List[int]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Report"
    os.makedirs("reports", exist_ok=True)
    report_path = f"reports/{task_id}_report.xlsx"


    # Set fixed column widths
    def set_column_widths(ws, widths):
        for i, width in enumerate(widths, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

# Apply once when setting up the worksheet
    set_column_widths(ws, [15, 20, 15, 15, 15, 20])  # Adjust widths as needed
 

    def add_section_title(title): 
        ws.append([title])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=7)
        cell.alignment = Alignment(horizontal="center", vertical="center", indent=0)  # center, no indent


    def add_table_header(headers):
        ws.append(headers)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", indent=1)  # add some left padding


    def add_row(data):
        ws.append(data)
        for col in range(1, len(data)+1):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(indent=1)  # left padding


    def add_summary_row(label, data):
        ws.append([label] + data)
        for col in range(1, len(data)+3):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(indent=1)

    ### INVENTORY SECTION ###
    add_section_title("Inventory Upload")

    ### PURCHASE SECTION ###
    add_section_title("Purchases")
    add_table_header(["Purchase ID", "Date", "Total Items", "Total Stock", "Total Cost", "Total VAT", "Grand Total"])
    
    

    purchase_data_with_items = (
    session.query(
        Purchase.id,
        Purchase.date,
        func.count(PurchaseItem.id).label("total_items"),
        func.sum(PurchaseItem.quantity).label("total_stock"),
        Purchase.total,
        Purchase.total_tax,
        Purchase.grand_total
        
     )
    .join(PurchaseItem, Purchase.id == PurchaseItem.purchase_id)
    .filter(Purchase.id.in_(purchase_ids))
    .group_by(Purchase.id)
    .order_by(Purchase.date)
    .all()
    )

    totalItems = 0
    totalStock = 0
    totalCost = 0   
    totalVAT = 0
    grandTotal = 0
    for purchase in purchase_data_with_items:
        totalItems += purchase[2] or 0
        totalStock += purchase[3] or 0
        totalCost += purchase[4] or 0   
        totalVAT += purchase[5] or 0
        grandTotal += purchase[6] or 0
        add_row([
            purchase[0],
            purchase[1],
            purchase[2],
            purchase[3],
            purchase[4],
            purchase[5],
            purchase[6]
        ])
       

    add_summary_row("Total Purchases", ['', totalItems, totalStock, totalCost, totalVAT, grandTotal])


    # Save the report
    wb.save(report_path)
    return report_path

